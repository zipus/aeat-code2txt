#!/usr/bin/env python3
"""
Convert an AEAT XLSX layout to one CSV per sheet.

Prefer xlsx2csv for better handling of merged cells and formatting.
Fallback to openpyxl if xlsx2csv is not available.
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path


def convert_with_xlsx2csv(src: Path, out_dir: Path) -> bool:
    try:
        from xlsx2csv import Xlsx2csv
    except Exception:
        return False
    import openpyxl

    out_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    converter = Xlsx2csv(
        str(src),
        outputencoding="utf-8",
        skipemptyrows=False,
    )
    for sheet_name in wb.sheetnames:
        name = sheet_name.strip().replace(" ", "_")
        out_path = out_dir / f"{name}.csv"
        converter.convert(str(out_path), sheetname=sheet_name)
    return True


def convert_with_openpyxl(src: Path, out_dir: Path) -> None:
    import openpyxl

    out_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(src, data_only=True)
    for ws in wb.worksheets:
        name = ws.title.strip().replace(" ", "_")
        out_path = out_dir / f"{name}.csv"
        with out_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(list(row))


def normalize_filenames(out_dir: Path) -> None:
    for path in out_dir.glob("*.csv"):
        new_name = path.name.replace(" ", "_")
        if new_name != path.name:
            path.rename(out_dir / new_name)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path, help="Input XLSX file")
    parser.add_argument("out_dir", type=Path, help="Output directory for CSVs")
    args = parser.parse_args()

    src = args.xlsx
    out_dir = args.out_dir

    if not src.exists():
        raise SystemExit(f"Input not found: {src}")

    used_xlsx2csv = convert_with_xlsx2csv(src, out_dir)
    if not used_xlsx2csv:
        convert_with_openpyxl(src, out_dir)

    normalize_filenames(out_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
